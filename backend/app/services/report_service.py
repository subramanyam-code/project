from datetime import date, timedelta
from typing import Optional
from sqlalchemy.orm import Session
from sqlalchemy import select, func, and_
import io, uuid

from app.models.daily_status import DailyStatus, TaskStatus
from app.models.user import User
from app.models.project import Project


class ReportService:
    def __init__(self, db: Session):
        self.db = db

    def _query(self, company_id, user_id, project_id, department_id, start, end):
        q = (select(DailyStatus).join(User, DailyStatus.user_id == User.id)
             .join(Project, DailyStatus.project_id == Project.id)
             .where(DailyStatus.submit_date >= start, DailyStatus.submit_date <= end))
        if company_id: q = q.where(User.company_id == company_id)
        if user_id: q = q.where(DailyStatus.user_id == user_id)
        if project_id: q = q.where(DailyStatus.project_id == project_id)
        if department_id: q = q.where(User.department_id == department_id)
        return q

    def _format(self, items, label):
        rows, total_hours, summary = [], 0.0, {"completed": 0, "in_progress": 0, "blocked": 0, "not_started": 0}
        for s in items:
            rows.append({"id": str(s.id), "user_name": s.user.full_name if s.user else "", "project_name": s.project.project_name if s.project else "",
                         "task_title": s.task_title, "status": s.status, "hours_worked": s.hours_worked, "blockers": s.blockers, "submit_date": str(s.submit_date)})
            total_hours += s.hours_worked or 0
            summary[s.status] = summary.get(s.status, 0) + 1
        return {"label": label, "total_entries": len(items), "total_hours": round(total_hours, 2), "summary": summary, "entries": rows}

    def daily_report(self, report_date, company_id=None, user_id=None, project_id=None, department_id=None):
        items = self.db.execute(self._query(company_id, user_id, project_id, department_id, report_date, report_date)).scalars().all()
        return self._format(items, str(report_date))

    def weekly_report(self, week_start, company_id=None, user_id=None, project_id=None, department_id=None):
        week_end = week_start + timedelta(days=6)
        items = self.db.execute(self._query(company_id, user_id, project_id, department_id, week_start, week_end)).scalars().all()
        return self._format(items, f"{week_start} – {week_end}")

    def monthly_report(self, year, month, company_id=None, user_id=None, project_id=None, department_id=None):
        start = date(year, month, 1)
        end = date(year + 1, 1, 1) - timedelta(days=1) if month == 12 else date(year, month + 1, 1) - timedelta(days=1)
        items = self.db.execute(self._query(company_id, user_id, project_id, department_id, start, end)).scalars().all()
        return self._format(items, f"{year}-{month:02d}")

    def custom_report(self, start_date, end_date, company_id=None, user_id=None, project_id=None, department_id=None):
        items = self.db.execute(self._query(company_id, user_id, project_id, department_id, start_date, end_date)).scalars().all()
        return self._format(items, f"{start_date} – {end_date}")

    def productivity_report(self, start_date, end_date, company_id=None, department_id=None):
        q = (select(User.id.label("user_id"), User.first_name, User.last_name,
                    func.count(DailyStatus.id).label("total"),
                    func.sum(DailyStatus.hours_worked).label("total_hours"),
                    func.count(DailyStatus.id).filter(DailyStatus.status == TaskStatus.COMPLETED).label("completed"),
                    func.count(DailyStatus.id).filter(DailyStatus.status == TaskStatus.BLOCKED).label("blocked"))
             .join(DailyStatus, DailyStatus.user_id == User.id, isouter=True)
             .where(and_(DailyStatus.submit_date >= start_date, DailyStatus.submit_date <= end_date))
             .group_by(User.id, User.first_name, User.last_name))
        if company_id: q = q.where(User.company_id == company_id)
        if department_id: q = q.where(User.department_id == department_id)
        rows = self.db.execute(q).all()
        return {"label": f"Productivity {start_date} – {end_date}", "employees": [
            {"user_id": str(r.user_id), "name": f"{r.first_name} {r.last_name}",
             "total_submissions": r.total or 0, "total_hours": round(float(r.total_hours or 0), 2),
             "completed": r.completed or 0, "blocked": r.blocked or 0,
             "completion_rate": round((r.completed or 0) / max(r.total or 1, 1) * 100, 1)} for r in rows]}

    def export_csv(self, data):
        import csv
        out = io.StringIO()
        if not data.get("entries"): return b""
        w = csv.DictWriter(out, fieldnames=data["entries"][0].keys())
        w.writeheader(); w.writerows(data["entries"])
        return out.getvalue().encode()

    def export_excel(self, data):
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"
        if data.get("entries"):
            headers = list(data["entries"][0].keys())
            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=c, value=h.replace("_", " ").title())
                cell.font = Font(color="FFFFFF", bold=True)
                cell.fill = PatternFill("solid", fgColor="1E40AF")
            for r, entry in enumerate(data["entries"], 2):
                for c, k in enumerate(headers, 1):
                    ws.cell(row=r, column=c, value=str(entry.get(k, "")))
        buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

    def export_pdf(self, data):
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        elements = [Paragraph(f"Report: {data.get('label', '')}", styles["Title"]), Spacer(1, 12)]
        if data.get("entries"):
            headers = ["User", "Project", "Task", "Status", "Hours", "Date"]
            rows = [headers] + [[e.get("user_name",""), e.get("project_name",""), e.get("task_title","")[:40],
                                  e.get("status",""), str(e.get("hours_worked",0)), e.get("submit_date","")] for e in data["entries"]]
            t = Table(rows, colWidths=[120,120,180,80,50,80])
            t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1E40AF")),
                                   ("TEXTCOLOR",(0,0),(-1,0),colors.white),("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
                                   ("GRID",(0,0),(-1,-1),0.5,colors.grey)]))
            elements.append(t)
        doc.build(elements)
        return buf.getvalue()
